import os
import openpyxl
from typing import Dict, Any, List
from app.schemas.parsed_document import ParsedDocument
from app.parsers.canonical.document import CanonicalDocument, Page, Block, Table, Coord, ImageInfo
from app.parsers.canonical.workbook import CanonicalWorkbook, WorkbookMetadata, Sheet, TableInfo, Cell

class CanonicalBuilder:
    @staticmethod
    def build_document(parsed: ParsedDocument) -> CanonicalDocument:
        """
        Converts a general ParsedDocument (PDF, PPTX, DOCX, Image) into a CanonicalDocument.
        """
        # Group content blocks by page index (or slide index)
        pages_dict: Dict[int, List[Any]] = {}
        for block in parsed.content:
            p_idx = block.page if block.page is not None else (block.slide if block.slide is not None else 0)
            if p_idx not in pages_dict:
                pages_dict[p_idx] = []
            pages_dict[p_idx].append(block)

        pages = []
        for p_idx in sorted(pages_dict.keys()):
            blocks = []
            tables = []
            images = []
            page_text_parts = []

            for block in pages_dict[p_idx]:
                if block.content_type == "text":
                    # Add standard block
                    blocks.append(Block(
                        block_id=block.id,
                        type=block.section.type if block.section.type != "unknown" else "text",
                        content=block.raw_text,
                        analysis_source="native",
                        coordinates=None  # Can be populated if coordinate mapping exists
                    ))
                    if block.raw_text:
                        page_text_parts.append(block.raw_text)
                elif block.content_type == "table":
                    tables.append(Table(
                        table_id=block.id,
                        grid=block.rows or [],
                        analysis_source="native",
                        coordinates=None
                    ))
                    # Fallback text representation of table
                    table_lines = ["\t".join([str(cell) for cell in row]) for row in (block.rows or [])]
                    page_text_parts.append("\n".join(table_lines))

            pages.append(Page(
                page_index=p_idx,
                dimensions=None,
                ocr_confidence=None,
                text="\n".join(page_text_parts),
                blocks=blocks,
                tables=tables,
                images=images
            ))

        return CanonicalDocument(
            document_id=parsed.document_id,
            document_type=parsed.document_type,
            metadata={
                "file_name": parsed.document_name,
                "file_size": parsed.metadata.file_size,
                "page_count": len(pages),
                "created_at": parsed.metadata.created_at
            },
            pages=pages
        )

    @staticmethod
    def build_workbook(parsed: ParsedDocument, file_path: str) -> CanonicalWorkbook:
        """
        Builds a CanonicalWorkbook from a ParsedDocument and the original Excel file path.
        Extracts spreadsheet metadata, hidden sheets, active sheets, cell relationships, and formula strings.
        """
        hidden_sheets = []
        active_sheet = None
        defined_names_list = []
        sheet_count = 0

        # Load metadata and formula cells using openpyxl
        sheets_data = {}
        if os.path.exists(file_path):
            try:
                wb = openpyxl.load_workbook(file_path, data_only=False)
                sheet_count = len(wb.sheetnames)
                active_sheet = wb.active.title if wb.active else None
                defined_names_list = [name for name in wb.defined_names.keys()] if wb.defined_names else []

                # Find hidden sheets
                for sheet in wb.worksheets:
                    if sheet.sheet_state == "hidden":
                        hidden_sheets.append(sheet.title)

                # Iterate and cache formula strings
                for sheet_name in wb.sheetnames:
                    ws = wb[sheet_name]
                    sheets_data[sheet_name] = {}
                    
                    # Inspect formulas and currencies
                    for r_idx in range(1, ws.max_row + 1):
                        for c_idx in range(1, ws.max_column + 1):
                            cell_obj = ws.cell(row=r_idx, column=c_idx)
                            cell_ref = cell_obj.coordinate
                            
                            # Cache values
                            cell_info = {
                                "formula": str(cell_obj.value) if cell_obj.value and str(cell_obj.value).startswith("=") else None,
                                "number_format": cell_obj.number_format,
                                "currency": "USD" if cell_obj.number_format and "$" in cell_obj.number_format else None
                            }
                            sheets_data[sheet_name][cell_ref] = cell_info
                wb.close()
            except Exception:
                pass

        # Build Sheet objects from ParsedDocument content
        sheets = []
        for block in parsed.content:
            if block.content_type != "table" or not block.sheet:
                continue

            sheet_name = block.sheet
            safe_sheet_data = sheets_data.get(sheet_name, {})

            # Standardize cell representations
            cells = []
            rows_grid = block.rows or []
            for r_idx, row_list in enumerate(rows_grid):
                for c_idx, val in enumerate(row_list):
                    # Coordinates in Excel are 1-indexed
                    excel_row = r_idx + 1
                    excel_col = c_idx + 1
                    
                    # Convert column index to Excel column string (e.g. 1 -> A, 2 -> B)
                    col_str = openpyxl.utils.get_column_letter(excel_col)
                    cell_ref = f"{col_str}{excel_row}"

                    # Look up parsed formula / format from cache
                    cell_meta = safe_sheet_data.get(cell_ref, {})
                    formula = cell_meta.get("formula")
                    number_format = cell_meta.get("number_format")
                    currency = cell_meta.get("currency")

                    cells.append(Cell(
                        row=excel_row,
                        col=excel_col,
                        reference=cell_ref,
                        value=str(val) if val is not None else None,
                        raw_value=val,
                        formula=formula,
                        number_format=number_format,
                        currency=currency
                    ))

            # Build simple sheet relationships mapping dependencies
            referenced_sheets = []
            dependencies_str = ""
            for cell in cells:
                if cell.formula:
                    # Detect if formula references other sheets
                    for other_sheet in parsed.metadata.statistics.sheets == len(sheets_data) and sheets_data.keys() or []:
                        if other_sheet != sheet_name and other_sheet in cell.formula:
                            if other_sheet not in referenced_sheets:
                                referenced_sheets.append(other_sheet)
                            dependencies_str += f"{other_sheet} -> {sheet_name}; "

            tables = [TableInfo(
                table_name=f"{sheet_name}_table",
                cells=cells
            )]

            sheets.append(Sheet(
                sheet_name=sheet_name,
                relationships={
                    "referenced_sheets": referenced_sheets,
                    "dependencies": dependencies_str.strip() or None
                },
                tables=tables
            ))

        workbook_metadata = WorkbookMetadata(
            sheet_count=sheet_count or len(sheets),
            defined_names=defined_names_list,
            hidden_sheets=hidden_sheets,
            active_sheet=active_sheet
        )

        return CanonicalWorkbook(
            document_id=parsed.document_id,
            metadata={
                "file_name": parsed.document_name,
                "sheet_names": [sheet.sheet_name for sheet in sheets]
            },
            workbook_metadata=workbook_metadata,
            sheets=sheets
        )
