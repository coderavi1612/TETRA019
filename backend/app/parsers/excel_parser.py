from openpyxl import load_workbook
import openpyxl
import os
import datetime

from app.schemas.parsed_document import ParsedDocument, ContentBlock, BlockSource, SectionInfo, DocumentMetadata, ParserInfo, DocumentStatistics
from app.parsers.base import BaseParser

class ExcelParser(BaseParser):
    def parse(self, file_path: str, company_id: str, document_type: str) -> ParsedDocument:
        doc_name = os.path.basename(file_path)
        file_size = os.path.getsize(file_path)
        from app.core import get_utc_now_iso
        created_at = get_utc_now_iso()
        
        wb = load_workbook(file_path, data_only=True)
        sheet_names = wb.sheetnames
        sheets_count = len(sheet_names)
        
        content_blocks = []
        sequence = 0
        tables_count = 0
        words_count = 0
        
        for sheet_name in sheet_names:
            ws = wb[sheet_name]
            
            rows = []
            sheet_words = 0
            for row in ws.iter_rows(values_only=True):
                # Filter out trailing completely empty rows to save space and maintain quality
                if not any(cell is not None for cell in row):
                    continue
                    
                row_data = []
                for cell in row:
                    if isinstance(cell, (datetime.datetime, datetime.date)):
                        val = cell.isoformat()
                    else:
                        val = cell
                    
                    if val is not None:
                        sheet_words += len(str(val).split())
                    row_data.append(val)
                rows.append(row_data)
                
            if not rows:
                continue
                
            words_count += sheet_words
            
            safe_sheet_name = "".join([c if c.isalnum() else "_" for c in sheet_name]).lower()
            block_id = f"{document_type}_sheet_{safe_sheet_name}"
            
            block = ContentBlock(
                id=block_id,
                sequence=sequence,
                content_type="table",
                sheet=sheet_name,
                raw_text="",
                rows=rows,
                section=SectionInfo(),
                source=BlockSource(file=doc_name, sheet=sheet_name)
            )
            content_blocks.append(block)
            sequence += 1
            tables_count += 1
            
        wb.close()
        
        statistics = DocumentStatistics(
            pages=0,
            slides=0,
            sheets=sheets_count,
            blocks=len(content_blocks),
            tables=tables_count,
            words=words_count
        )
        
        metadata = DocumentMetadata(
            company_id=company_id,
            file_size=file_size,
            extension=".xlsx",
            mime_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            created_at=created_at,
            parser=ParserInfo(name="openpyxl", version=openpyxl.__version__),
            statistics=statistics
        )
        
        document_id = f"{company_id}_{document_type}"
        
        return ParsedDocument(
            document_id=document_id,
            document_name=doc_name,
            document_type=document_type,
            status="parsed",
            metadata=metadata,
            content=content_blocks,
            warnings=[],
            errors=[]
        )
